#!/usr/bin/env python3
"""
camp_closing_report.py — hourly CLOSING report, formatted like the EC2
`simple_camp_db.py` active-camps-by-portal export.

It REUSES camp_closing's data engine (same snapshot DB, same success-rate
lookups at 1.6 and 2.1, same verdicts) — it does NOT pause anything. It just
presents the live picture as three portal tabs (SM / SML / NBP), each split
into sections:

    ACTIVE · >=50% spent (decision-ready)   <- the closing decisions live here
    ACTIVE · <50% spent  (too early)
    LEARNING                                 (leave alone)
    PAUSED / CLOSED

Columns match simple_camp_db.py plus the two success targets already in the
pipeline:
  date · account · campaign · created · learning · status · budget · spend ·
  spend% · ROAS · ROAS last 3h · remaining ₹ · ROAS needed @1.6 ·
  Success @1.6 · Success @2.1 · verdict · campaign_id

created_time comes from the snapshot table (no extra call); learning status is
one adsets read per account (Meta pixel data). Reads the same DB the closing
engine uses.

Usage:
  META_ACCESS_TOKEN=... python3 scripts/v2/camp_closing_report.py --db state/camp_snapshots.db
  ... --xlsx out/camp_closing_report.xlsx      # also write a styled .xlsx
  ... --no-sheet                               # skip the Google Sheet write
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from camp_closing import (SCHEMA, build_first_activity, collect,  # noqa: E402
                          _fmt_rate)
from success_lookup import build_both, TARGET_ROAS, TARGET_ROAS_2  # noqa: E402

IST = timezone(timedelta(hours=5, minutes=30))
API = 'https://graph.facebook.com/v19.0'
REPO = Path(__file__).resolve().parent.parent.parent
PORTALS = ('SM', 'SML', 'NBP')

HEADER = ['date', 'account', 'campaign', 'created', 'learning', 'status',
          'budget ₹', 'spend ₹', 'spend %', 'ROAS', 'ROAS last 3h',
          'remaining ₹', f'ROAS needed @{TARGET_ROAS}',
          f'Success @{TARGET_ROAS}', f'Success @{TARGET_ROAS_2}',
          'verdict', 'campaign_id']


# ── learning status (one adsets read per account, aggregated to campaign) ──
def _get(url):
    for i in range(4):
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 503) and i < 3:
                import time
                time.sleep(5 * (i + 1)); continue
            return {}
        except Exception:
            return {}
    return {}


def configured_accounts() -> list[str]:
    out = []
    for name in ('config/accounts.env', '.env'):
        p = REPO / name
        if p.exists():
            for line in p.read_text(errors='ignore').splitlines():
                m = re.match(r'^\s*[A-Z][A-Z0-9_]*\s*=\s*(act_\d+)\s*$', line.strip())
                if m and m.group(1) not in out:
                    out.append(m.group(1))
    return out


def _learn(effective_status: str, lsi: dict) -> str:
    ls = str((lsi or {}).get('status') or '').upper()
    if ls == 'LEARNING':
        return 'Learning'
    if ls == 'LEARNING_LIMITED':
        return 'Learning Limited'
    return 'Active'


def fetch_learning(token: str, accounts: list[str]) -> dict:
    """{campaign_id: 'Learning' | 'Learning Limited' | 'Active'} from active adsets."""
    out: dict[str, str] = {}
    rank = {'Learning': 2, 'Learning Limited': 1, 'Active': 0}
    for aid in accounts:
        url = (f'{API}/{aid}/adsets?' + urllib.parse.urlencode({
            'fields': 'campaign_id,effective_status,learning_stage_info',
            'effective_status': json.dumps(['ACTIVE']),
            'limit': 500, 'access_token': token}))
        while url:
            j = _get(url)
            for s in j.get('data', []):
                cid = s.get('campaign_id')
                if not cid:
                    continue
                lv = _learn(s.get('effective_status'), s.get('learning_stage_info'))
                if rank[lv] > rank.get(out.get(cid, 'Active'), 0):
                    out[cid] = lv
            url = (j.get('paging') or {}).get('next')
    return out


def created_map(con, day: str) -> dict:
    """{campaign_id: created_time[:10]} from the latest snapshot slot of the day."""
    latest = con.execute(
        "SELECT MAX(hour_slot) FROM campaign_hourly_snapshots WHERE hour_slot LIKE ?",
        (day + '%',)).fetchone()[0]
    if not latest:
        return {}
    out = {}
    for cid, ct in con.execute(
            "SELECT campaign_id, created_time FROM campaign_hourly_snapshots WHERE hour_slot=?",
            (latest,)):
        out[cid] = (ct or '')[:10]
    return out


def roas_needed(budget: float, spend: float, revenue: float, target: float) -> str:
    rem = budget - spend
    if rem <= 0:
        return ''
    need = (target * budget - revenue) / rem
    return round(need if need > 0 else 0, 2)


def section_of(r: dict) -> str:
    if r['learning'] in ('Learning', 'Learning Limited'):
        return 'LEARNING'
    if r['status'] != 'Active':
        return 'PAUSED / CLOSED'
    return 'ACTIVE_HI' if r['spend_pct'] >= 50 else 'ACTIVE_LO'


SECTIONS = [
    ('ACTIVE_HI', '🔴 ACTIVE · ≥50% spent — decision-ready'),
    ('ACTIVE_LO', '🟢 ACTIVE · <50% spent — too early to judge'),
    ('LEARNING', '🟡 LEARNING — leave alone'),
    ('PAUSED / CLOSED', '⚪ PAUSED / CLOSED'),
]


def row_values(r: dict) -> list:
    return [
        r['slot'][:10], r['account_name'], r['campaign_name'], r['created'], r['learning'],
        r['status'], round(r['daily_budget']), round(r['spend']), r['spend_pct'], r['roas'],
        r['roas_3h'] if r['roas_3h'] is not None else '',
        round(r['daily_budget'] - r['spend']),
        roas_needed(r['daily_budget'], r['spend'], r['revenue'], TARGET_ROAS),
        _fmt_rate(r['success_rate'], r['sr_fallback']),
        _fmt_rate(r['success_rate_21'], r['sr21_fallback']),
        r['verdict'], r['campaign_id'],
    ]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', default='state/camp_snapshots.db')
    ap.add_argument('--day', default=None)
    ap.add_argument('--sheet', default=os.environ.get(
        'LIVE_ROAS_SHEET_ID', '1eW2_qPdsKJ8zAV5-hsXA5HtfVH9NwDhQLyHYGKz5hXk'))
    ap.add_argument('--sa', default=os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE',
                                                   'google-service-account.json'))
    ap.add_argument('--no-sheet', action='store_true')
    ap.add_argument('--xlsx', default=None)
    args = ap.parse_args()

    now = datetime.now(IST)
    day = args.day or now.strftime('%Y-%m-%d')
    token = os.environ.get('META_ACCESS_TOKEN', '')

    if not Path(args.db).exists():
        print(f'FATAL: missing DB {args.db}'); sys.exit(1)

    lk16, lk21 = build_both(args.db, exclude_day=day)
    con = sqlite3.connect(args.db)
    con.executescript(SCHEMA)
    fa = build_first_activity(con)
    rows = collect(con, day, lk16, lk21, fa)
    if not rows:
        print(f'no snapshots for {day} — nothing to report'); return

    created = created_map(con, day)
    learning = fetch_learning(token, configured_accounts()) if token else {}
    if not token:
        print('  (no META_ACCESS_TOKEN — learning column blank)')
    for r in rows:
        r['created'] = created.get(r['campaign_id'], '')
        r['learning'] = learning.get(r['campaign_id'], 'Active' if r['status'] == 'Active' else '')

    # bucket per portal → section
    by_portal = {p: {s[0]: [] for s in SECTIONS} for p in PORTALS}
    for r in rows:
        p = r['portal']
        if p in by_portal:
            by_portal[p][section_of(r)].append(r)
    for p in PORTALS:
        for sec in by_portal[p]:
            key = (1 if sec.startswith('ACTIVE') else 0)  # decision sections: worst ROAS first
            by_portal[p][sec].sort(key=lambda r: (r['roas'] if key else -r['spend']))

    for p in PORTALS:
        counts = {s[0]: len(by_portal[p][s[0]]) for s in SECTIONS}
        print(f"{p}: " + " · ".join(f"{lbl.split('—')[0].strip()[:18]} {counts[k]}" for k, lbl in SECTIONS))

    if not args.no_sheet:
        write_sheet(args.sheet, args.sa, by_portal, day, rows[0]['slot'], lk16.n_camp_days)
    if args.xlsx:
        write_xlsx(args.xlsx, by_portal, day, rows[0]['slot'])


# ── Google Sheet: 3 portal tabs, sectioned, camp_daily_db-style colours ──
def _cs(color):
    return {'red': color[0] / 255, 'green': color[1] / 255, 'blue': color[2] / 255}


def write_sheet(sheet_id, sa_file, by_portal, day, slot, n_camp_days):
    import gspread
    gc = gspread.service_account(filename=sa_file)
    sh = gc.open_by_key(sheet_id)
    for p in PORTALS:
        tab = f'Closing {p}'
        # recreate the tab each run so old conditional-format rules don't pile up
        try:
            sh.del_worksheet(sh.worksheet(tab))
        except gspread.WorksheetNotFound:
            pass
        total = sum(len(v) for v in by_portal[p].values())
        ws = sh.add_worksheet(title=tab, rows=total + 40, cols=len(HEADER) + 1)
        title = (f'🛑 CLOSING · {p} · {day} · slot {slot[-5:]} IST · {total} campaigns · '
                 f"'Success @{TARGET_ROAS}/{TARGET_ROAS_2}' = % of past camp-days in the same "
                 f'spend%×ROAS band that finished at/above target ({n_camp_days} camp-days) · '
                 f"'ROAS needed @{TARGET_ROAS}' = ROAS required on remaining budget to hit "
                 f'{TARGET_ROAS} blended by EOD · Meta pixel')
        block = [[title], HEADER]
        section_rows = []          # 0-based sheet row indices of section headers
        for key, label in SECTIONS:
            items = by_portal[p][key]
            if not items:
                continue
            section_rows.append(len(block))
            block.append([f'{label}  ({len(items)})'] + [''] * (len(HEADER) - 1))
            block.extend(row_values(r) for r in items)
        ws.update(range_name='A1', values=block)

        sid = ws.id
        reqs = []
        # header + title bold, section rows bold on a grey band
        def fmt(r0, r1, cell):
            reqs.append({'repeatCell': {
                'range': {'sheetId': sid, 'startRowIndex': r0, 'endRowIndex': r1,
                          'startColumnIndex': 0, 'endColumnIndex': len(HEADER)},
                'cell': {'userEnteredFormat': cell}, 'fields': 'userEnteredFormat'}})
        fmt(1, 2, {'backgroundColor': _cs((31, 78, 120)),
                   'textFormat': {'bold': True, 'foregroundColor': _cs((255, 255, 255))}})
        for rr in section_rows:
            fmt(rr, rr + 1, {'backgroundColor': _cs((226, 232, 240)),
                             'textFormat': {'bold': True}})
        # colour scales on ROAS (col 9), spend% (col 8), ROAS-needed (col 12)
        last = len(block)

        def scale(col, lo, mid, hi, lo_v, mid_v, hi_v, reverse=False):
            a, c = ((246, 105, 107), (99, 190, 123)) if not reverse else ((99, 190, 123), (246, 105, 107))
            reqs.append({'addConditionalFormatRule': {'index': 0, 'rule': {
                'ranges': [{'sheetId': sid, 'startRowIndex': 2, 'endRowIndex': last,
                            'startColumnIndex': col, 'endColumnIndex': col + 1}],
                'gradientRule': {
                    'minpoint': {'color': _cs(a), 'type': 'NUMBER', 'value': str(lo_v)},
                    'midpoint': {'color': _cs((255, 235, 132)), 'type': 'NUMBER', 'value': str(mid_v)},
                    'maxpoint': {'color': _cs(c), 'type': 'NUMBER', 'value': str(hi_v)}}}}})
        scale(9, None, None, None, 0, 1.5, 3)                      # ROAS: red→green
        scale(8, None, None, None, 0, 100, 200, reverse=True)     # spend%: high=red
        scale(12, None, None, None, 0, 1.6, 4, reverse=True)      # ROAS needed: high=hard=red
        reqs.append({'updateSheetProperties': {
            'properties': {'sheetId': sid, 'gridProperties': {'frozenRowCount': 2}},
            'fields': 'gridProperties.frozenRowCount'}})
        try:
            sh.batch_update({'requests': reqs})
        except Exception as e:
            print(f'  ({p} formatting skipped: {str(e)[:80]})')
        print(f"wrote 'Closing {p}' ({total} rows)")


def write_xlsx(path, by_portal, day, slot):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    wb = Workbook(); wb.remove(wb.active)
    HF = PatternFill('solid', fgColor='1F4E78'); HFT = Font(bold=True, color='FFFFFF', size=10)
    SEC = PatternFill('solid', fgColor='E2E8F0'); SECF = Font(bold=True)
    for p in PORTALS:
        ws = wb.create_sheet(p)
        ws.append(HEADER)
        for c in ws[1]:
            c.fill = HF; c.font = HFT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for key, label in SECTIONS:
            items = by_portal[p][key]
            if not items:
                continue
            ws.append([f'{label}  ({len(items)})'])
            sr = ws.max_row
            for cc in ws[sr]:
                cc.fill = SEC; cc.font = SECF
            for r in items:
                ws.append(row_values(r))
        n = ws.max_row
        if n > 1:
            ws.conditional_formatting.add(f'J2:J{n}', ColorScaleRule(
                start_type='num', start_value=0, start_color='F8696B',
                mid_type='num', mid_value=1.5, mid_color='FFEB84',
                end_type='num', end_value=3, end_color='63BE7B'))
            ws.conditional_formatting.add(f'I2:I{n}', ColorScaleRule(
                start_type='num', start_value=0, start_color='DDEBF7',
                mid_type='num', mid_value=100, mid_color='FFEB84',
                end_type='num', end_value=200, end_color='F8696B'))
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADER))}1'
        for col in ws.columns:
            try:
                mx = max(len(str(c.value or '')) for c in col)
            except Exception:
                mx = 14
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 2, 48)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f'wrote xlsx {path}')


if __name__ == '__main__':
    main()
